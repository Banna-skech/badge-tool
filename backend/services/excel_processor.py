"""
Excel employee table processing service.
Parses raw HR system exports and generates formatted badge/seat-card Excel files.

Target format (per 0804工牌信息1.xlsx):
  Sheet "工牌":   姓名, 英文名/姓名拼音, 工号  (sorted by 工号)
  Sheet "座位牌": 姓名, 英文名/姓名拼音, 工号, 二级部门, 详细职位名称, 工作地点  (sorted by 部门)
"""
import io
import re
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypinyin import pinyin, Style


# ── Comprehensive Chinese Surname → Pinyin Dictionary ──────────────────
# Covers all common Chinese surnames, with special handling for polyphonic
# characters where pypinyin defaults to the wrong reading for surnames.

SURNAME_PINYIN: Dict[str, str] = {
    # ── Polyphonic surnames (pypinyin gets these WRONG) ──
    '乐': 'Yue',       # pypinyin→'Le',  surname→Yue
    '单': 'Shan',      # pypinyin→'Dan', surname→Shan
    '解': 'Xie',       # pypinyin→'Jie', surname→Xie
    '仇': 'Qiu',       # pypinyin→'Chou', surname→Qiu
    '查': 'Zha',       # pypinyin→'Cha', surname→Zha
    '翟': 'Zhai',      # pypinyin→'Di', surname→Zhai
    '盖': 'Ge',        # pypinyin→'Gai', surname→Ge
    '区': 'Ou',        # pypinyin→'Qu', surname→Ou
    '覃': 'Qin',       # pypinyin→'Tan', surname→Qin
    '曾': 'Zeng',      # pypinyin→'Ceng', surname→Zeng
    '尉': 'Yu',        # pypinyin→'Wei', surname→Yu (as surname)
    '种': 'Chong',     # pypinyin→'Zhong', surname→Chong
    '员': 'Yun',       # pypinyin→'Yuan', surname→Yun
    '祭': 'Zhai',      # pypinyin→'Ji', surname→Zhai
    '繁': 'Po',        # pypinyin→'Fan', surname→Po
    '缪': 'Miao',      # pypinyin→'Mou', surname→Miao
    '能': 'Nai',       # pypinyin→'Neng', surname→Nai
    '句': 'Gou',       # pypinyin→'Ju', surname→Gou
    '阿': 'E',         # pypinyin→'A', surname→E (as surname)
    '隗': 'Wei',       # pypinyin→'Kui', surname→Wei
    '秘': 'Bi',        # pypinyin→'Mi', surname→Bi
    '卜': 'Bu',        # pypinyin→'Bo', surname→Bu
    '召': 'Shao',      # pypinyin→'Zhao', surname→Shao
    '折': 'She',       # pypinyin→'Zhe', surname→She
    '郇': 'Xun',       # pypinyin→'Huan', surname→Xun (also Huan)
    '殳': 'Shu',       # pypinyin→'Shu'
    '卞': 'Bian',      # pypinyin→'Bian'
    '乜': 'Nie',       # pypinyin→'Mie', surname→Nie
    '宓': 'Fu',        # pypinyin→'Mi', surname→Fu
    '禚': 'Zhuo',      # pypinyin→'Zhuo'
    '郗': 'Xi',        # pypinyin→'Xi' (also Chi)
    '郜': 'Gao',       # pypinyin→'Gao'
    '钭': 'Tou',       # pypinyin→'Dou', surname→Tou
    '剌': 'La',        # pypinyin→'La'
    '驷': 'Si',        # pypinyin→'Si'
    '酆': 'Feng',      # pypinyin→'Feng'
    '夔': 'Kui',       # pypinyin→'Kui'
    '厍': 'She',       # pypinyin→'She'
    '訾': 'Zi',        # pypinyin→'Zi'
    '阚': 'Kan',       # pypinyin→'Kan'
    '逯': 'Lu',        # pypinyin→'Lu'

    # ── Compound surnames (pypinyin splits them wrongly) ──
    '万俟': 'Moqi',
    '尉迟': 'Yuchi',
    '长孙': 'Zhangsun',
    '司马': 'Sima',
    '上官': 'Shangguan',
    '欧阳': 'Ouyang',
    '夏侯': 'Xiahou',
    '诸葛': 'Zhuge',
    '闻人': 'Wenren',
    '东方': 'Dongfang',
    '赫连': 'Helian',
    '皇甫': 'Huangfu',
    '公羊': 'Gongyang',
    '澹台': 'Tantai',
    '公冶': 'Gongye',
    '宗政': 'Zongzheng',
    '濮阳': 'Puyang',
    '淳于': 'Chunyu',
    '单于': 'Chanyu',
    '太叔': 'Taishu',
    '申屠': 'Shentu',
    '公孙': 'Gongsun',
    '仲孙': 'Zhongsun',
    '轩辕': 'Xuanyuan',
    '令狐': 'Linghu',
    '锺离': 'Zhongli',
    '宇文': 'Yuwen',
    '慕容': 'Murong',
    '鲜于': 'Xianyu',
    '闾丘': 'Lvqiu',
    '司徒': 'Situ',
    '司空': 'Sikong',
    '丌官': 'Qiguan',
    '司寇': 'Sikou',
    '子车': 'Ziche',
    '颛孙': 'Zhuansun',
    '端木': 'Duanmu',
    '巫马': 'Wuma',
    '公西': 'Gongxi',
    '漆雕': 'Qidiao',
    '乐正': 'Yuezheng',
    '壤驷': 'Rangsi',
    '公良': 'Gongliang',
    '拓拔': 'Tuoba',
    '夹谷': 'Jiagu',
    '宰父': 'Zaifu',
    '谷梁': 'Guliang',
    '段干': 'Duangan',
    '百里': 'Baili',
    '东郭': 'Dongguo',
    '南门': 'Nanmen',
    '呼延': 'Huyan',
    '羊舌': 'Yangshe',
    '微生': 'Weisheng',
    '梁丘': 'Liangqiu',
    '左丘': 'Zuoqiu',
    '东门': 'Dongmen',
    '西门': 'Ximen',
    '南宫': 'Nangong',
    '第五': 'Diwu',
}

# ── Column auto-detection ────────────────────────────────────────────

# Keywords to match source columns → target fields
COLUMN_KEYWORDS: Dict[str, List[str]] = {
    "姓名":     ["姓名", "名字", "员工姓名", "Name"],
    "英文名":   ["英文名", "英文姓名", "姓名拼音", "拼音", "EngName", "English"],
    "工号":     ["工号", "员工编号", "JobNumber", "UserID"],
    "二级部门": ["二级部门", "部门", "ThirdLevelOrganization"],
    "三级部门": ["三级部门", "子部门", "FourthLevelOrganization"],
    "职位":     ["详细职位名称", "职位名称", "职位", "岗位", "职务"],
    "工作地点": ["工作地点", "工作地", "Location", "WorkPlace"],
}

TARGET_FIELDS = ["姓名", "英文名", "工号", "二级部门", "三级部门", "职位", "工作地点"]


def _match_column(col_name: str, keywords: List[str]) -> bool:
    """Check if a column name matches any of the given keywords."""
    clean = col_name.strip().lower()
    for kw in keywords:
        if kw.lower() in clean:
            return True
    return False


def auto_detect_columns(chinese_headers: List[str]) -> Dict[str, Optional[int]]:
    """
    Given a list of Chinese column headers, return a mapping:
        {target_field: column_index (0-based), or None if not found}
    """
    mapping: Dict[str, Optional[int]] = {}
    for field in TARGET_FIELDS:
        mapping[field] = None
        keywords = COLUMN_KEYWORDS.get(field, [])
        for i, h in enumerate(chinese_headers):
            if _match_column(h, keywords):
                mapping[field] = i
                break
    return mapping


# ── Parsing ───────────────────────────────────────────────────────────

def _clean_header_row(row: tuple) -> List[str]:
    """Convert a header row tuple to cleaned string list."""
    return [str(c).strip() if c is not None else "" for c in row]


def parse_raw_employee_excel(file_bytes: bytes) -> Dict[str, Any]:
    """
    Parse a raw employee Excel export.

    Handles the typical 2-row header format:
      Row 1: system internal field IDs (e.g. "LookupPrefix_UserID_JobNumber")
      Row 2: Chinese display names (e.g. "工号")
      Row 3+: data

    Returns:
        {
            "chinese_headers": ["姓名", "英文名/姓名拼音", ...],
            "system_headers": ["Name", "EngName", ...],
            "column_map": {"姓名": 0, "英文名": 1, ...},
            "rows": [ [...], [...], ... ],   # data rows (string values)
            "total_rows": 20,
            "preview": [ {...}, {...} ],      # first 5 rows as dicts
        }
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    raw_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not raw_rows:
        return {
            "chinese_headers": [], "system_headers": [],
            "column_map": {}, "rows": [],
            "total_rows": 0, "preview": [],
        }

    # Detect header structure
    row0 = _clean_header_row(raw_rows[0])
    row1 = _clean_header_row(raw_rows[1]) if len(raw_rows) > 1 else [""] * len(row0)

    # Determine which row is the Chinese header, which is the system header
    # System headers often contain UUIDs, "LookupPrefix", or internal field IDs
    system_like = any(
        "lookup" in c.lower() or "userid" in c.lower() or
        "ext" in c.lower() or "-" in c
        for c in row0 if c
    )
    # Check row0 for CJK → single-header format (Chinese headers only)
    row0_has_cjk = any('一' <= ch <= '鿿' for c in row0 for ch in c)
    # Check row1 for CJK → two-header format (system IDs + Chinese)
    row1_has_cjk = any('一' <= ch <= '鿿' for c in row1 for ch in c) if row1 else False

    if system_like and row1_has_cjk:
        # Two-row header: Row 0 = system IDs, Row 1 = Chinese names
        system_headers = row0
        chinese_headers = row1
        data_start = 2
    elif row0_has_cjk:
        # Single header row (Chinese)
        system_headers = [""] * len(row0)
        chinese_headers = row0
        data_start = 1
    else:
        # Fallback: treat row 0 as headers
        system_headers = [""] * len(row0)
        chinese_headers = row0
        data_start = 1

    # Auto-detect column mapping
    column_map = auto_detect_columns(chinese_headers)

    # Extract data rows
    data_rows = []
    for row in raw_rows[data_start:]:
        vals = [str(c).strip() if c is not None else "" for c in row]
        # Skip completely empty rows
        if any(v for v in vals):
            data_rows.append(vals)

    # Generate preview (first 5 rows as dicts with target fields)
    preview = []
    for row in data_rows[:5]:
        item = {}
        for field, idx in column_map.items():
            if idx is not None and idx < len(row):
                item[field] = row[idx]
            else:
                item[field] = ""
        preview.append(item)

    return {
        "chinese_headers": chinese_headers,
        "system_headers": system_headers,
        "column_map": {k: v for k, v in column_map.items() if v is not None},
        "column_map_all": column_map,  # includes None values
        "rows": data_rows,
        "total_rows": len(data_rows),
        "preview": preview,
    }


# ── Output generation ─────────────────────────────────────────────────

def _natural_sort_key(value: str) -> Tuple:
    """Sort key for natural sorting of strings with numbers."""
    parts = re.split(r'(\d+)', value)
    result = []
    for p in parts:
        if p.isdigit():
            result.append((0, int(p)))
        else:
            result.append((1, p.lower()))
    return tuple(result)


def _get_field(row: List[str], column_map: Dict[str, int], field: str) -> str:
    """Safely get a field value from a data row."""
    idx = column_map.get(field)
    if idx is not None and idx < len(row):
        return row[idx]
    return ""


def _normalize_engname(engname: str) -> str:
    """Capitalize the first letter of each part of the English name."""
    if not engname:
        return ""
    parts = re.split(r'([ _])', engname)
    result = []
    for part in parts:
        if part in (' ', '_'):
            result.append(part)
        elif part:
            result.append(part[0].upper() + part[1:].lower() if part else "")
    return ''.join(result)


def _format_badge_engname(engname: str) -> str:
    """
    For 工牌: use short English name (first part before space/underscore),
    with first letter capitalized.
    "wee" → "Wee"
    "Lan Songqi" → "Lan"
    """
    if not engname:
        return ""
    normalized = _normalize_engname(engname)
    # If there's a space, take the first word
    if ' ' in normalized:
        return normalized.split(' ')[0]
    # If there's an underscore, take the first part
    if '_' in normalized:
        return normalized.split('_')[0]
    return normalized


def _get_surname_pinyin(chinese_name: str) -> str:
    """
    Extract Chinese surname and convert to pinyin with capitalized first letter.

    Strategy:
    1. Try single-char lookup in SURNAME_PINYIN dict (polyphonic surnames take priority)
    2. Try compound surname (2-char) lookup if the single-char is NOT in the dict
       (this avoids rare compound surnames like 乐正 from overriding common polyphonic
       single-char surnames like 乐→Yue)
    3. Fall back to pypinyin library
    4. Worst case: return the original character (should rarely happen)
    """
    if not chinese_name:
        return ""
    surname_char = chinese_name[0]

    # Step 1: single-char polyphonic surname — highest priority
    if surname_char in SURNAME_PINYIN:
        return SURNAME_PINYIN[surname_char]

    # Step 2: compound surname — only if single-char is NOT a known polyphonic
    if len(chinese_name) >= 2:
        compound = chinese_name[:2]
        if compound in SURNAME_PINYIN:
            return SURNAME_PINYIN[compound]

    # Step 3: fall back to pypinyin
    py = pinyin(surname_char, style=Style.NORMAL, heteronym=False)
    if py and py[0]:
        result = py[0][0]
        # If pypinyin returns Chinese characters (unrecognized), use original char
        if result and not any('一' <= ch <= '鿿' for ch in result):
            return result.capitalize()

    # Step 4: worst case fallback
    return surname_char


def _format_seat_engname(engname: str, chinese_name: str) -> str:
    """
    For 座位牌: use "EnglishFirstName_SurnamePinyin" format,
    with English first name capitalized.
    e.g. 王瑜 (Florence) → Florence_Wang
         贺晓薇 (wee) → Wee_He
         兰颂琦 (Lan Songqi) → Lan_Lan
    """
    if not engname:
        return ""
    normalized = _normalize_engname(engname)
    # Extract English first name
    if ' ' in normalized:
        first_name = normalized.split(' ')[0]
    elif '_' in normalized:
        first_name = normalized.split('_')[0]
    else:
        first_name = normalized

    # Get surname pinyin from Chinese name
    surname_pinyin = _get_surname_pinyin(chinese_name)

    if surname_pinyin:
        return f"{first_name}_{surname_pinyin}"
    return first_name


def _build_dept(row: List[str], column_map: Dict[str, int]) -> str:
    """
    Build the department string for 座位牌.
    Format: "二级部门 - 三级部门"
    If 三级部门 is empty, just use 二级部门.
    """
    dept2 = _get_field(row, column_map, "二级部门")
    dept3 = _get_field(row, column_map, "三级部门")

    if dept2 and dept3:
        return f"{dept2} - {dept3}"
    elif dept2:
        return dept2
    elif dept3:
        return dept3
    return ""


def _auto_column_width(ws, min_width: int = 8, max_width: int = 40):
    """Set column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col_cells:
            if col_letter is None:
                col_letter = get_column_letter(cell.column)
            if cell.value:
                # CJK characters are wider — count them as ~2.2
                val = str(cell.value)
                char_len = 0
                for ch in val:
                    if '一' <= ch <= '鿿' or '㐀' <= ch <= '䶿':
                        char_len += 2.2
                    else:
                        char_len += 1.1
                max_len = max(max_len, char_len)
        if col_letter:
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def generate_output_excel(
    parsed_data: Dict[str, Any],
    output_path: Optional[str] = None,
) -> bytes:
    """
    Generate the formatted output Excel with two sheets.

    Args:
        parsed_data: result from parse_raw_employee_excel()
        output_path: if given, save to this path; always returns bytes

    Returns:
        Excel file as bytes
    """
    column_map = parsed_data["column_map"]
    data_rows = parsed_data["rows"]

    # ── Build data records ──
    records = []
    for row in data_rows:
        engname_raw = _get_field(row, column_map, "英文名")
        chinese_name = _get_field(row, column_map, "姓名")
        dept = _build_dept(row, column_map)

        rec = {
            "姓名": chinese_name,
            "工号": _get_field(row, column_map, "工号"),
            "英文名_raw": engname_raw,
            "工牌_英文名": _format_badge_engname(engname_raw),
            "座位牌_英文名": _format_seat_engname(engname_raw, chinese_name),
            "二级部门": dept,
            "详细职位名称": _get_field(row, column_map, "职位"),
            "工作地点": _get_field(row, column_map, "工作地点"),
        }
        records.append(rec)

    # ── Create workbook ──
    wb = Workbook()

    # Style definitions
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def write_header(ws, headers: List[str]):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def write_data_rows(ws, records: List[Dict], fields: List[str], start_row: int = 2):
        for row_idx, rec in enumerate(records, start_row):
            for col_idx, field in enumerate(fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=rec.get(field, ""))
                cell.border = thin_border

    # ── Sheet 1: 工牌 (sorted by 工号) ──
    ws_badge = wb.active
    ws_badge.title = "工牌"

    badge_fields = ["姓名", "工牌_英文名", "工号"]
    badge_headers = ["姓名", "英文名/姓名拼音", "工号"]
    badge_records = sorted(records, key=lambda r: _natural_sort_key(r["工号"]))

    write_header(ws_badge, badge_headers)
    write_data_rows(ws_badge, badge_records, badge_fields)
    _auto_column_width(ws_badge)

    # ── Sheet 2: 座位牌 (sorted by 部门) ──
    ws_seat = wb.create_sheet("座位牌")

    seat_fields = ["姓名", "座位牌_英文名", "工号", "二级部门", "详细职位名称", "工作地点"]
    seat_headers = ["姓名", "英文名/姓名拼音", "工号", "二级部门", "详细职位名称", "工作地点"]
    seat_records = sorted(records, key=lambda r: _natural_sort_key(r["二级部门"]))

    write_header(ws_seat, seat_headers)
    write_data_rows(ws_seat, seat_records, seat_fields)
    _auto_column_width(ws_seat)

    # ── Save ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    result_bytes = output.read()

    if output_path:
        wb.save(output_path)

    wb.close()
    return result_bytes


# ── Convenience: single-call pipeline ─────────────────────────────────

def process_employee_excel(file_bytes: bytes) -> bytes:
    """
    Single-call pipeline: parse raw Excel → generate formatted output.
    Returns the processed Excel as bytes.
    """
    parsed = parse_raw_employee_excel(file_bytes)
    return generate_output_excel(parsed)
