"""
Name list parsing service.
Supports Excel (.xlsx) and text (.txt) formats.
Auto-detects name columns by scanning for Chinese characters.
"""
import io
import re
from typing import List, Dict, Optional, Tuple

from openpyxl import load_workbook


# CJK Unified Ideographs range + common extensions
CJK_RANGES = [
    (0x4E00, 0x9FFF),   # CJK Unified Ideographs
    (0x3400, 0x4DBF),   # CJK Unified Ideographs Extension A
    (0xF900, 0xFAFF),   # CJK Compatibility Ideographs
    (0x2F800, 0x2FA1F), # CJK Compatibility Ideographs Supplement
]


def _is_chinese_char(ch: str) -> bool:
    """Check if a character is within CJK ranges (likely a Chinese character)."""
    cp = ord(ch)
    for start, end in CJK_RANGES:
        if start <= cp <= end:
            return True
    return False


def _chinese_ratio(text: str) -> float:
    """Calculate the proportion of Chinese characters in a string."""
    if not text:
        return 0.0
    chinese_count = sum(1 for ch in text if _is_chinese_char(ch))
    return chinese_count / len(text)


def parse_excel(file_bytes: bytes) -> Dict:
    """
    Parse an Excel (.xlsx) name list.

    Returns:
        {
            "columns": ["列A", "列B", ...],
            "data": {"列A": [...], "列B": [...], ...},
            "suggested_name_column": "姓名",  # or None
            "total_rows": 10
        }
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"columns": [], "data": {}, "suggested_name_column": None, "total_rows": 0}

    # Detect header row (first row with non-empty cells)
    header_row = rows[0]
    # Clean header names
    columns = [str(h).strip() if h else "" for h in header_row]

    # If headers are empty, generate column names
    if all(not c for c in columns):
        columns = [f"列{chr(65+i)}" for i in range(len(header_row))]

    # Extract data rows
    data_rows = rows[1:] if len(rows) > 1 else []

    # Build data dict
    data: Dict[str, List] = {col: [] for col in columns}
    for row in data_rows:
        for i, col in enumerate(columns):
            val = row[i] if i < len(row) else None
            data[col].append(str(val).strip() if val is not None else "")

    # Auto-detect name column
    suggested = _detect_name_column(data, columns)

    wb.close()

    return {
        "columns": columns,
        "data": data,
        "suggested_name_column": suggested,
        "total_rows": len(data_rows),
    }


def _detect_name_column(data: Dict[str, List], columns: List[str]) -> Optional[str]:
    """
    Auto-detect which column contains person names.
    Strategy:
    1. Look for columns with Chinese name patterns (2-4 chars, high CJK ratio)
    2. Prefer columns with header like "姓名", "名字", "员工", "Name"
    """
    NAME_KEYWORDS = ["姓名", "名字", "员工", "姓名", "名称", "name", "姓名（"]

    best_col = None
    best_score = 0.0

    for col in columns:
        values = [v for v in data.get(col, []) if v]  # non-empty
        if not values:
            continue

        score = 0.0

        # Check header keyword match
        col_lower = col.lower().strip()
        for kw in NAME_KEYWORDS:
            if kw.lower() in col_lower:
                score += 2.0
                break

        # Check values: Chinese names are usually 2-4 characters, high CJK ratio
        cjk_scores = []
        for val in values[:20]:  # sample first 20
            if not val:
                continue
            ratio = _chinese_ratio(val)
            length = len(val)
            # Ideal name: 2-4 chars, high Chinese ratio
            if 2 <= length <= 4 and ratio > 0.8:
                cjk_scores.append(1.0)
            elif 1 <= length <= 6 and ratio > 0.5:
                cjk_scores.append(0.6)
            elif length > 6 and ratio > 0.3:
                cjk_scores.append(0.3)
            else:
                cjk_scores.append(0.0)

        if cjk_scores:
            avg_score = sum(cjk_scores) / len(cjk_scores)
            coverage = len([s for s in cjk_scores if s > 0]) / len(cjk_scores)
            score += avg_score * coverage * 3.0

        if score > best_score:
            best_score = score
            best_col = col

    return best_col if best_score > 1.0 else None


def parse_text(file_bytes: bytes) -> List[str]:
    """
    Parse a text (.txt) name list.
    Each line = one name. Empty lines are skipped.
    """
    text = file_bytes.decode('utf-8-sig')  # handle BOM
    lines = text.strip().split('\n')
    names = [line.strip() for line in lines if line.strip()]
    return names


def match_photos_to_names(
    sorted_filenames: List[str],
    name_list: List[str]
) -> Tuple[List[Tuple[str, str]], Optional[str]]:
    """
    Match photos (sorted alphabetically) with name list by position.

    Args:
        sorted_filenames: List of original filenames, sorted alphabetically
        name_list: List of names in order

    Returns:
        Tuple of:
        - List of (filename, assigned_name) tuples
        - Warning message string (or None if all good)
    """
    warning = None
    pairs = []

    if len(sorted_filenames) != len(name_list):
        warning = f"照片数量 ({len(sorted_filenames)}) 与名单人数 ({len(name_list)}) 不一致"

    for i, fname in enumerate(sorted_filenames):
        if i < len(name_list):
            pairs.append((fname, name_list[i]))
        else:
            # Extra photos beyond name list — use original filename
            base = fname.rsplit('.', 1)[0] if '.' in fname else fname
            pairs.append((fname, base))

    return pairs, warning


def extract_names_from_column(data: Dict[str, List], column: str) -> List[str]:
    """Extract names from a specific column in the parsed Excel data."""
    return [v for v in data.get(column, []) if v]  # filter empty
